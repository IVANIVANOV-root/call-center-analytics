"""Call Center Analyzer — Flask веб-сервер."""
import os
import io
import json
import functools
from datetime import datetime
from flask import (Flask, render_template, jsonify, request,
                   Response, session, redirect, send_file)
from werkzeug.utils import secure_filename
from config import PORT, AUDIO_DIR
import db
from processor import scan_and_process, get_active_count

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "callcontrol-secret-2026")
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024

ALLOWED_EXTS = {".mp3", ".wav", ".ogg", ".m4a", ".aac", ".flac", ".opus"}
PRICE_DIR = os.environ.get("PRICE_DIR", "/app/price")

DEFAULT_REPORT_PROMPT = (
    "Ты аналитик качества колл-центра. Дай краткие стратегические рекомендации руководителю.\n\n"
    "Данные:\n{data}\n\n"
    "Составь 4 рекомендации. Каждая: заголовок (до 5 слов) и текст (1-2 предложения).\n"
    "Ответь только JSON: [{\"title\":\"...\",\"text\":\"...\"}]"
)


# ─── Auth helpers ─────────────────────────────────────────────────────────────

def login_required(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized"}), 401
            return redirect("/")
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        if session.get("role") != "admin":
            return jsonify({"error": "forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper


# ─── Pages ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


# ─── Health ───────────────────────────────────────────────────────────────────

@app.route("/api/health")
def api_health():
    return jsonify({"status": "ok", "processing": get_active_count()})


# ─── Auth API ─────────────────────────────────────────────────────────────────

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.json or {}
    user = db.authenticate(data.get("username", ""), data.get("password", ""))
    if not user:
        return jsonify({"error": "Неверный логин или пароль"}), 401
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["role"] = user["role"]
    return jsonify({
        "ok": True,
        "username": user["username"],
        "role": user["role"],
        "theme": user.get("theme", "light"),
        "has_gigachat_token": bool(user.get("gigachat_token")),
    })


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/me")
def api_me():
    if session.get("user_id"):
        user = db.get_user_by_id(session["user_id"])
        return jsonify({
            "authenticated": True,
            "username": session["username"],
            "role": session["role"],
            "theme": user.get("theme", "light") if user else "light",
            "has_gigachat_token": bool(user.get("gigachat_token")) if user else False,
        })
    return jsonify({"authenticated": False})


@app.route("/api/me/token", methods=["GET", "PUT"])
@login_required
def api_me_token():
    if request.method == "GET":
        token = db.get_user_gigachat_token(session["user_id"])
        return jsonify({"has_token": bool(token), "token_preview": token[:8] + "…" if len(token) > 8 else ""})
    data = request.json or {}
    token = data.get("token", "").strip()
    db.set_user_gigachat_token(session["user_id"], token)
    return jsonify({"ok": True})



@app.route("/api/me/whisper-prompt", methods=["GET", "PUT"])
@login_required
def api_me_whisper_prompt():
    if request.method == "GET":
        prompt = db.get_user_whisper_prompt(session["user_id"])
        return jsonify({"prompt": prompt})
    data = request.json or {}
    db.set_user_whisper_prompt(session["user_id"], data.get("prompt", "").strip())
    return jsonify({"ok": True})


@app.route("/api/settings/report-prompt", methods=["GET", "PUT"])
@login_required
def api_report_prompt():
    if request.method == "GET":
        content = db.get_setting("report_prompt") or DEFAULT_REPORT_PROMPT
        return jsonify({"content": content})
    if session.get("role") != "admin":
        return jsonify({"error": "forbidden"}), 403
    data = request.json or {}
    db.set_setting("report_prompt", data.get("content", "").strip())
    return jsonify({"ok": True})


@app.route("/api/settings/report-prompt/reset", methods=["POST"])
@login_required
def api_report_prompt_reset():
    if session.get("role") != "admin":
        return jsonify({"error": "forbidden"}), 403
    db.set_setting("report_prompt", "")
    return jsonify({"ok": True})


@app.route("/api/me/theme", methods=["PUT"])
@login_required
def api_me_theme():
    data = request.json or {}
    theme = data.get("theme", "light")
    if theme not in ("light", "dark"):
        theme = "light"
    db.set_user_theme(session["user_id"], theme)
    return jsonify({"ok": True})


# ─── User management (admin only) ─────────────────────────────────────────────

@app.route("/api/users")
@login_required
@admin_required
def api_users():
    return jsonify(db.get_all_users())


@app.route("/api/users", methods=["POST"])
@login_required
@admin_required
def api_create_user():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "viewer")
    if not username or not password:
        return jsonify({"error": "Логин и пароль обязательны"}), 400
    if role not in ("admin", "viewer"):
        return jsonify({"error": "Недопустимая роль"}), 400
    user = db.create_user(username, password, role)
    if not user:
        return jsonify({"error": "Пользователь уже существует"}), 409
    return jsonify({"id": user["id"], "username": user["username"], "role": user["role"]})


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@login_required
@admin_required
def api_delete_user(user_id):
    if user_id == session["user_id"]:
        return jsonify({"error": "Нельзя удалить себя"}), 400
    db.delete_user(user_id)
    return jsonify({"ok": True})


@app.route("/api/users/<int:user_id>/password", methods=["PUT"])
@login_required
@admin_required
def api_change_password(user_id):
    data = request.json or {}
    password = data.get("password", "")
    if len(password) < 3:
        return jsonify({"error": "Пароль слишком короткий"}), 400
    db.change_password(user_id, password)
    return jsonify({"ok": True})


# ─── Stats & Calls ────────────────────────────────────────────────────────────

@app.route("/api/stats")
@login_required
def api_stats():
    date_from = request.args.get("date_from")
    date_to   = request.args.get("date_to")
    operator  = request.args.get("operator") or None
    stats = db.get_stats(session["user_id"], date_from=date_from, date_to=date_to, operator=operator)
    stats["processing"] = get_active_count()
    return jsonify(stats)


@app.route("/api/stats/periods")
@login_required
def api_stats_periods():
    operator = request.args.get("operator") or None
    return jsonify(db.get_period_stats(session["user_id"], operator=operator))


@app.route("/api/calls")
@login_required
def api_calls():
    date_from = request.args.get("date_from")
    date_to   = request.args.get("date_to")
    operator  = request.args.get("operator") or None
    result    = request.args.get("result") or None
    search    = request.args.get("search") or None
    sort_by   = request.args.get("sort_by") or None
    sort_dir  = request.args.get("sort_dir", "desc")
    return jsonify(db.get_all_calls(
        session["user_id"],
        date_from=date_from, date_to=date_to,
        operator=operator, result=result,
        search=search, sort_by=sort_by, sort_dir=sort_dir,
    ))


@app.route("/api/operators")
@login_required
def api_operators():
    return jsonify(db.get_operators_list(session["user_id"]))


@app.route("/api/process", methods=["POST"])
@login_required
def api_process():
    count = scan_and_process()
    return jsonify({"ok": True, "queued": count})


@app.route("/api/calls/<int:call_id>/cancel", methods=["POST"])
@login_required
def api_cancel_call(call_id):
    row = db.get_call(call_id, session["user_id"])
    if not row:
        return jsonify({"error": "not found"}), 404
    db.set_cancelled(call_id)
    return jsonify({"ok": True})


@app.route("/api/cancel-pending", methods=["POST"])
@login_required
@admin_required
def api_cancel_pending():
    count = db.cancel_all_pending(session["user_id"])
    return jsonify({"ok": True, "cancelled": count})


@app.route("/api/reprocess/<int:call_id>", methods=["POST"])
@login_required
def api_reprocess(call_id):
    row = db.get_call(call_id, session["user_id"])
    if not row:
        return jsonify({"error": "not found"}), 404
    db.set_status(call_id, "pending")
    db.clear_operator_if_auto(call_id)
    from processor import _process_one, _executor
    _executor.submit(_process_one, call_id)
    return jsonify({"ok": True})


@app.route("/api/reprocess-selected", methods=["POST"])
@login_required
def api_reprocess_selected():
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids or not isinstance(ids, list):
        return jsonify({"error": "ids required"}), 400
    from processor import _process_one, _executor
    count = 0
    for call_id in ids:
        row = db.get_call(call_id, session["user_id"])
        if row:
            db.set_status(call_id, "pending")
            db.clear_operator_if_auto(call_id)
            _executor.submit(_process_one, call_id)
            count += 1
    return jsonify({"ok": True, "queued": count})


# ─── Upload ───────────────────────────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    from processor import _parse_filename
    from transcriber import get_audio_duration
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "no files"}), 400

    operator = request.form.get("operator", "").strip() or None
    upload_batch = datetime.now().strftime("%Y-%m-%d %H:%M")
    saved, skipped = [], []
    user_id = session["user_id"]

    for f in files:
        fname = secure_filename(f.filename)
        if not fname:
            continue
        ext = os.path.splitext(fname)[1].lower()
        if ext not in ALLOWED_EXTS:
            skipped.append(fname)
            continue
        dest = os.path.join(AUDIO_DIR, fname)
        if os.path.exists(dest):
            skipped.append(fname + " (уже существует)")
            continue
        f.save(dest)
        call_date, call_time = _parse_filename(fname)
        duration = get_audio_duration(dest)
        db.upsert_call(fname, call_date, call_time,
                       user_id=user_id, upload_batch=upload_batch,
                       operator=operator, duration=duration)
        saved.append(fname)

    queued = scan_and_process() if saved else 0
    return jsonify({"ok": True, "saved": saved, "skipped": skipped, "queued": queued,
                    "batch": upload_batch})


# ─── Call updates ─────────────────────────────────────────────────────────────

@app.route("/api/calls/<int:call_id>/operator", methods=["PUT"])
@login_required
def api_update_operator(call_id):
    data = request.json or {}
    operator = data.get("operator", "").strip() or None
    db.update_call_operator(call_id, operator, session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/calls/<int:call_id>/tags", methods=["GET", "PUT"])
@login_required
def api_call_tags(call_id):
    if request.method == "GET":
        row = db.get_call(call_id, session["user_id"])
        if not row:
            return jsonify({"error": "not found"}), 404
        return jsonify({"tags": row.get("tags", [])})
    data = request.json or {}
    tags = data.get("tags", [])
    db.update_call_tags(call_id, tags, session["user_id"])
    return jsonify({"ok": True})


# ─── Delete ───────────────────────────────────────────────────────────────────

@app.route("/api/calls/<int:call_id>", methods=["DELETE"])
@login_required
def api_delete_call(call_id):
    filename = db.delete_call(call_id, session["user_id"])
    if not filename:
        return jsonify({"error": "not found"}), 404
    audio_path = os.path.join(AUDIO_DIR, filename)
    if os.path.exists(audio_path):
        os.remove(audio_path)
    return jsonify({"ok": True})


@app.route("/api/calls", methods=["DELETE"])
@login_required
def api_delete_all_calls():
    filenames = db.delete_all_calls(session["user_id"])
    for fname in filenames:
        path = os.path.join(AUDIO_DIR, fname)
        if os.path.exists(path):
            os.remove(path)
    return jsonify({"ok": True, "deleted": len(filenames)})


@app.route("/api/calls/batch", methods=["DELETE"])
@login_required
def api_delete_batch():
    batch = request.args.get("batch") or (request.json or {}).get("batch", "")
    if not batch:
        return jsonify({"error": "batch required"}), 400
    filenames = db.delete_calls_by_batch(batch, session["user_id"])
    for fname in filenames:
        path = os.path.join(AUDIO_DIR, fname)
        if os.path.exists(path):
            os.remove(path)
    return jsonify({"ok": True, "deleted": len(filenames)})


# ─── Export Excel ─────────────────────────────────────────────────────────────

@app.route("/api/export")
@login_required
def api_export():
    import math
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_from = request.args.get("date_from")
    date_to   = request.args.get("date_to")
    operator  = request.args.get("operator") or None
    result    = request.args.get("result") or None
    search    = request.args.get("search") or None

    calls = db.get_all_calls(
        session["user_id"],
        date_from=date_from, date_to=date_to,
        operator=operator, result=result, search=search,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Звонки"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="E0E7FF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Дата", "Время", "Оператор", "Длительность",
        "Итог", "Услуга", "Оценка", "Тон", "Качество",
        "Клиент", "Резюме", "Ошибки", "Рекомендации", "Теги"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 24

    result_colors = {
        "записан": "D1FAE5", "не записан": "FEE2E2",
        "недозвон": "FEF3C7", "перезвон": "FEF3C7",
        "справка": "EEF2FF", "прочее": "F1F5F9",
    }

    for row_idx, call in enumerate(calls, 2):
        errors = call.get("operator_errors", [])
        errors_str = "; ".join(errors) if errors else ""
        tags_str = ", ".join(call.get("tags", []))
        dur = call.get("duration")
        dur_str = f"{dur // 60}:{dur % 60:02d}" if dur else ""

        row_data = [
            call.get("call_date", ""),
            call.get("call_time", ""),
            call.get("operator", ""),
            dur_str,
            call.get("result", ""),
            call.get("service", ""),
            call.get("score", ""),
            call.get("tone", ""),
            call.get("call_quality", ""),
            call.get("patient_name", ""),
            call.get("summary", ""),
            errors_str,
            call.get("recommendations", ""),
            tags_str,
        ]

        result_val = call.get("result", "")
        row_fill = PatternFill("solid", fgColor=result_colors.get(result_val, "FFFFFF"))

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (11, 12, 13)))
            cell.border = border
            if col == 5 and result_val:
                cell.fill = row_fill

        # Авто-высота по длинным текстовым колонкам (Резюме=11, Ошибки=12, Рекомендации=13)
        wrap_cols = {11: 40, 12: 40, 13: 40}  # col_idx: ширина колонки в символах
        max_lines = 1
        for col_idx, col_width in wrap_cols.items():
            val = row_data[col_idx - 1]
            if val:
                lines = sum(
                    max(1, math.ceil(len(line) / col_width))
                    for line in str(val).split("\n")
                )
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row_idx].height = max(18, max_lines * 15)

    col_widths = [12, 10, 18, 12, 14, 20, 9, 12, 14, 18, 40, 40, 40, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"calls_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── PDF Report helpers ───────────────────────────────────────────────────────

def _fmt_dur(sec):
    if not sec:
        return "—"
    return f"{sec // 60}:{sec % 60:02d}"


def _fmt_date_ru(d):
    if not d:
        return "—"
    try:
        return datetime.strptime(d, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return d


def _calc_report_stats(calls):
    from collections import Counter
    total = len(calls)
    scored = [c for c in calls if c.get("score") is not None]
    avg_score = round(sum(c["score"] for c in scored) / len(scored), 1) if scored else None
    durlist = [c["duration"] for c in calls if c.get("duration")]
    avg_dur = int(sum(durlist) / len(durlist)) if durlist else None

    results = {}
    for c in calls:
        r = (c.get("result") or "прочее").lower()
        results[r] = results.get(r, 0) + 1

    recorded = results.get("записан", 0)
    conversion = round(recorded / total * 100) if total else 0

    op_map = {}
    for c in calls:
        op = c.get("operator") or "—"
        if op not in op_map:
            op_map[op] = {"calls": 0, "scores": [], "recorded": 0, "durs": [], "errors": []}
        d = op_map[op]
        d["calls"] += 1
        if c.get("score") is not None:
            d["scores"].append(c["score"])
        if (c.get("result") or "").lower() == "записан":
            d["recorded"] += 1
        if c.get("duration"):
            d["durs"].append(c["duration"])
        d["errors"].extend(c.get("operator_errors") or [])

    for op, d in op_map.items():
        d["avg_score"] = round(sum(d["scores"]) / len(d["scores"]), 1) if d["scores"] else None
        d["conversion"] = round(d["recorded"] / d["calls"] * 100) if d["calls"] else 0
        d["avg_dur"] = int(sum(d["durs"]) / len(d["durs"])) if d["durs"] else None
        d["top_errors"] = [e for e, _ in Counter(d["errors"]).most_common(3)]

    score_bands = {"9–10": 0, "7–8": 0, "5–6": 0, "1–4": 0}
    for c in calls:
        s = c.get("score")
        if s is None:
            continue
        if s >= 9:
            score_bands["9–10"] += 1
        elif s >= 7:
            score_bands["7–8"] += 1
        elif s >= 5:
            score_bands["5–6"] += 1
        else:
            score_bands["1–4"] += 1

    return {
        "total": total, "avg_score": avg_score, "avg_dur": avg_dur,
        "conversion": conversion, "recorded": recorded,
        "operators": op_map, "results": results, "score_bands": score_bands,
    }


def _generate_ai_recommendations(calls, gigachat_token=None):
    import re
    from collections import Counter
    import requests as req
    from config import OLLAMA_BASE_URL, OLLAMA_MODEL

    op_data = {}
    # Берём не более 40 последних звонков чтобы уложиться в таймаут
    for c in calls[-40:]:
        op = c.get("operator") or "Неизвестный"
        if op not in op_data:
            op_data[op] = {"errors": [], "recs": [], "n": 0}
        op_data[op]["n"] += 1
        op_data[op]["errors"].extend(c.get("operator_errors") or [])
        if c.get("recommendations"):
            op_data[op]["recs"].append(c["recommendations"])

    lines = []
    for op, d in op_data.items():
        lines.append(f"Оператор: {op} (звонков: {d['n']})")
        if d["errors"]:
            ec = Counter(d["errors"])
            lines.append("  Ошибки: " + "; ".join(f"{e} (x{n})" for e, n in ec.most_common(5)))
        if d["recs"]:
            lines.append("  Рекомендации: " + " | ".join(d["recs"][:3]))
        lines.append("")

    tpl = db.get_setting("report_prompt") or DEFAULT_REPORT_PROMPT
    prompt = tpl.replace("{data}", "\n".join(lines))

    # GigaChat приоритет если есть токен
    if gigachat_token:
        try:
            from transcriber import _get_gigachat_token, GIGACHAT_BASE_URL, GIGACHAT_MODEL
            from config import GIGACHAT_BASE_URL as GC_URL, GIGACHAT_MODEL as GC_MODEL
            import urllib3, uuid
            urllib3.disable_warnings()
            token = _get_gigachat_token(gigachat_token)
            resp = req.post(
                f"{GC_URL}/chat/completions",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json={"model": GC_MODEL, "messages": [{"role": "user", "content": prompt}],
                      "temperature": 0.1, "max_tokens": 1000},
                verify=False, timeout=50,
            )
            resp.raise_for_status()
            raw = resp.json()["choices"][0]["message"]["content"].strip()
            match = re.search(r"\[.*\]", raw, re.DOTALL)
            if match:
                return json.loads(match.group())
        except Exception:
            pass  # fallback Ollama

    try:
        resp = req.post(
            f"{OLLAMA_BASE_URL}/api/generate",
            json={"model": OLLAMA_MODEL, "prompt": prompt, "stream": False},
            timeout=50,
        )
        raw = resp.json().get("response", "")
        raw = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL).strip()
        match = re.search(r"\[.*\]", raw, re.DOTALL)
        if match:
            return json.loads(match.group())
    except Exception:
        pass
    return []


def _svg_hbars(items, total, color_map, width=340):
    items = sorted(items, key=lambda x: -x[1])
    row_h = 32
    h = len(items) * row_h + 8
    bar_x, bar_end = 110, width - 70
    bar_w = bar_end - bar_x
    parts = [f'<svg width="{width}" height="{h}" viewBox="0 0 {width} {h}" xmlns="http://www.w3.org/2000/svg">']
    for i, (label, count) in enumerate(items):
        y = i * row_h
        pct = count / total if total else 0
        fill_w = max(0, int(bar_w * pct))
        color = color_map.get(label, "#94A3B8")
        parts.append(
            f'<text x="0" y="{y+20}" font-size="11" font-family="Arial,sans-serif" fill="#334155">{label}</text>'
            f'<rect x="{bar_x}" y="{y+7}" width="{bar_w}" height="16" rx="4" fill="#F1F5F9"/>'
        )
        if fill_w > 0:
            parts.append(f'<rect x="{bar_x}" y="{y+7}" width="{fill_w}" height="16" rx="4" fill="{color}"/>')
        parts.append(
            f'<text x="{bar_x + bar_w + 6}" y="{y+20}" font-size="11" font-family="Arial,sans-serif" fill="#64748B">'
            f'{count} ({int(pct*100)}%)</text>'
        )
    parts.append("</svg>")
    return "".join(parts)


def _build_report_html(calls, stats, ai_recs, date_from, date_to, operators_filter, include_summaries):
    PRIMARY = "#4F46E5"
    SUCCESS = "#16A34A"
    WARNING = "#D97706"
    DANGER  = "#DC2626"
    NEUTRAL = "#64748B"

    d_from = _fmt_date_ru(date_from) if date_from else "начало"
    d_to   = _fmt_date_ru(date_to)   if date_to   else "конец"
    generated = datetime.now().strftime("%d.%m.%Y %H:%M")
    ops_label = ", ".join(operators_filter) if operators_filter else "Все операторы"

    # ── inline SVG иконки (Bootstrap Icons paths) ──
    ICO_PHONE  = '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" fill="currentColor" viewBox="0 0 16 16"><path d="M3.654 1.328a.678.678 0 0 0-1.015-.063L1.605 2.3c-.483.484-.661 1.169-.45 1.77a17.6 17.6 0 0 0 4.168 6.608 17.6 17.6 0 0 0 6.608 4.168c.601.211 1.286.033 1.77-.45l1.034-1.034a.678.678 0 0 0-.063-1.015l-2.307-1.794a.68.68 0 0 0-.58-.122l-2.19.547a1.75 1.75 0 0 1-1.657-.459L5.482 8.062a1.75 1.75 0 0 1-.46-1.657l.548-2.19a.68.68 0 0 0-.122-.58z"/></svg>'
    ICO_CHECK  = '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" fill="currentColor" viewBox="0 0 16 16"><path d="M16 8A8 8 0 1 1 0 8a8 8 0 0 1 16 0m-3.97-3.03a.75.75 0 0 0-1.08.022L7.477 9.417 5.384 7.323a.75.75 0 0 0-1.06 1.06L6.97 11.03a.75.75 0 0 0 1.079-.02l3.992-4.99a.75.75 0 0 0-.01-1.05z"/></svg>'
    ICO_STAR   = '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" fill="currentColor" viewBox="0 0 16 16"><path d="M3.612 15.443c-.386.198-.824-.149-.746-.592l.83-4.73L.173 6.765c-.329-.314-.158-.888.283-.95l4.898-.696L7.538.792c.197-.39.73-.39.927 0l2.184 4.327 4.898.696c.441.062.612.636.282.95l-3.522 3.356.83 4.73c.078.443-.36.79-.746.592L8 13.187l-4.389 2.256z"/></svg>'
    ICO_CLOCK  = '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" fill="currentColor" viewBox="0 0 16 16"><path d="M8 3.5a.5.5 0 0 0-1 0V9a.5.5 0 0 0 .252.434l3.5 2a.5.5 0 0 0 .496-.868L8 8.71z"/><path d="M8 16A8 8 0 1 0 8 0a8 8 0 0 0 0 16m7-8A7 7 0 1 1 1 8a7 7 0 0 1 14 0"/></svg>'
    ICO_PERSON = '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" fill="currentColor" viewBox="0 0 16 16"><path d="M8 8a3 3 0 1 0 0-6 3 3 0 0 0 0 6m2-3a2 2 0 1 1-4 0 2 2 0 0 1 4 0m4 8c0 1-1 1-1 1H3s-1 0-1-1 1-4 6-4 6 3 6 4m-1-.004c-.001-.246-.154-.986-.832-1.664C11.516 10.68 10.289 10 8 10s-3.516.68-4.168 1.332c-.678.678-.83 1.418-.832 1.664z"/></svg>'
    ICO_BULB   = '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" fill="currentColor" viewBox="0 0 16 16"><path d="M2 6a6 6 0 1 1 10.174 4.31c-.203.196-.359.4-.453.619l-.762 1.769A.5.5 0 0 1 10.5 13a.5.5 0 0 1 0 1 .5.5 0 0 1 0 1l-.224.447a1 1 0 0 1-.894.553H6.618a1 1 0 0 1-.894-.553L5.5 15a.5.5 0 0 1 0-1 .5.5 0 0 1 0-1 .5.5 0 0 1-.46-.302l-.761-1.77a2 2 0 0 0-.453-.618A5.98 5.98 0 0 1 2 6m6-5a5 5 0 0 0-3.479 8.592c.263.254.514.564.676.941L5.83 12h4.342l.632-1.467c.162-.377.413-.687.676-.941A5 5 0 0 0 8 1"/></svg>'
    ICO_LIST   = '<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20" fill="currentColor" viewBox="0 0 16 16"><path fill-rule="evenodd" d="M2.5 12a.5.5 0 0 1 .5-.5h10a.5.5 0 0 1 0 1H3a.5.5 0 0 1-.5-.5m0-4a.5.5 0 0 1 .5-.5h10a.5.5 0 0 1 0 1H3a.5.5 0 0 1-.5-.5m0-4a.5.5 0 0 1 .5-.5h10a.5.5 0 0 1 0 1H3a.5.5 0 0 1-.5-.5"/></svg>'

    # WeasyPrint не наследует CSS color в SVG fill="currentColor" — подставляем цвет явно
    def ci(svg, color):
        return svg.replace('fill="currentColor"', f'fill="{color}"')

    def kpi(ico, label, val, sub, color):
        sub_h = f'<div class="kpi-sub">{sub}</div>' if sub else ""
        return (
            f'<div class="kpi-card">'
            f'<div class="kpi-icon" style="background:{color}1a">{ci(ico, color)}</div>'
            f'<div class="kpi-val" style="color:{color}">{val}</div>'
            f'<div class="kpi-lbl">{label}</div>{sub_h}</div>'
        )

    # ── Диаграммы ──
    result_colors = {
        "записан": SUCCESS, "не записан": DANGER,
        "недозвон": WARNING, "перезвон": WARNING,
        "справка": PRIMARY, "прочее": NEUTRAL,
    }
    score_colors = {"9–10": SUCCESS, "7–8": "#22C55E", "5–6": WARNING, "1–4": DANGER}
    outcome_svg = _svg_hbars(list(stats["results"].items()), stats["total"], result_colors)
    score_svg   = _svg_hbars(list(stats["score_bands"].items()), stats["total"], score_colors)

    # ── Таблица операторов ──
    op_rows = ""
    for op_name, od in sorted(stats["operators"].items(), key=lambda x: -x[1]["calls"]):
        initials = "".join(w[0].upper() for w in op_name.split()[:2]) if op_name != "—" else "?"
        sc = od["avg_score"]
        sc_color = SUCCESS if sc and sc >= 8 else WARNING if sc and sc >= 6 else DANGER
        cv_color = SUCCESS if od["conversion"] >= 50 else WARNING if od["conversion"] >= 30 else DANGER
        top_e = "; ".join(od["top_errors"][:2]) if od["top_errors"] else "—"
        op_rows += (
            f'<tr>'
            f'<td><div class="op-cell">'
            f'<div class="op-av">{initials}</div><span>{op_name}</span></div></td>'
            f'<td class="tc">{od["calls"]}</td>'
            f'<td class="tc" style="color:{sc_color};font-weight:600">{sc if sc is not None else "—"}</td>'
            f'<td class="tc" style="color:{cv_color};font-weight:600">{od["conversion"]}%</td>'
            f'<td class="tc">{_fmt_dur(od["avg_dur"])}</td>'
            f'<td class="sm">{top_e}</td>'
            f'</tr>'
        )

    # ── Строки таблицы звонков ──
    RBADGE = {
        "записан":    ("D1FAE5","065F46"), "не записан": ("FEE2E2","991B1B"),
        "недозвон":   ("FEF3C7","92400E"), "перезвон":   ("FEF3C7","92400E"),
        "справка":    ("EEF2FF","3730A3"), "прочее":     ("F1F5F9","475569"),
    }
    call_rows = ""
    for c in calls:
        res = (c.get("result") or "").lower()
        bg, fg = RBADGE.get(res, ("FFFFFF","374151"))
        sc  = c.get("score")
        sc_col = SUCCESS if sc and sc >= 8 else WARNING if sc and sc >= 6 else DANGER if sc else NEUTRAL
        call_rows += (
            f'<tr>'
            f'<td class="sm">{c.get("call_date","") or "—"}<br>'
            f'<span style="color:#94A3B8">{c.get("call_time","") or ""}</span></td>'
            f'<td>{c.get("operator","—") or "—"}</td>'
            f'<td class="tc">{_fmt_dur(c.get("duration"))}</td>'
            f'<td><span class="badge" style="background:#{bg};color:#{fg}">'
            f'{c.get("result","—") or "—"}</span></td>'
            f'<td>{c.get("service","—") or "—"}</td>'
            f'<td class="tc" style="color:{sc_col};font-weight:600">{sc or "—"}</td>'
            f'<td class="sm">{c.get("tone","—") or "—"}</td>'
            f'<td class="sm">{c.get("call_quality","—") or "—"}</td>'
            f'</tr>'
        )

    # ── Детализация по звонкам ──
    summaries_html = ""
    if include_summaries:
        cards = []
        for c in calls:
            if not (c.get("summary") or c.get("operator_errors") or c.get("recommendations")):
                continue
            parts_inner = []
            if c.get("summary"):
                parts_inner.append(
                    f'<div class="det-sec"><div class="det-lbl">Резюме</div>'
                    f'<div class="det-txt">{c["summary"]}</div></div>'
                )
            errs = c.get("operator_errors") or []
            if errs:
                items_li = "".join(f"<li>{e}</li>" for e in errs)
                parts_inner.append(
                    f'<div class="det-sec"><div class="det-lbl">Ошибки оператора</div>'
                    f'<ul class="err-ul">{items_li}</ul></div>'
                )
            if c.get("recommendations"):
                parts_inner.append(
                    f'<div class="det-sec"><div class="det-lbl">Рекомендации</div>'
                    f'<div class="det-txt">{c["recommendations"]}</div></div>'
                )
            cards.append(
                f'<div class="det-card">'
                f'<div class="det-hdr">'
                f'<span class="det-date">{c.get("call_date","") or ""} {c.get("call_time","") or ""}</span>'
                f'<span class="det-op">{c.get("operator","—") or "—"}</span>'
                f'<span class="det-svc">{c.get("service","") or ""}</span>'
                f'</div>{"".join(parts_inner)}</div>'
            )
        if cards:
            summaries_html = (
                f'<div class="section pb">'
                f'<div class="sec-hdr"><div class="sec-ico" style="background:{PRIMARY}1a">'
                f'{ci(ICO_LIST, PRIMARY)}</div><h2>Детализация по звонкам</h2></div>'
                + "".join(cards) + "</div>"
            )

    # ── AI-рекомендации ──
    ai_html = ""
    if ai_recs:
        rec_cards = "".join(
            f'<div class="rec-card">'
            f'<div class="rec-title">{r.get("title","")}</div>'
            f'<div class="rec-txt">{r.get("text","")}</div>'
            f'</div>'
            for r in (ai_recs or [])
        )
        if rec_cards:
            ai_html = (
                f'<div class="section">'
                f'<div class="sec-hdr"><div class="sec-ico" style="background:{PRIMARY}1a">'
                f'{ci(ICO_BULB, PRIMARY)}</div><h2>AI-рекомендации</h2>'
                f'<span class="sec-sub">Сформированы на основе {stats["total"]} звонков</span></div>'
                f'<div class="recs-col">{rec_cards}</div></div>'
            )

    # ── Раздел по операторам ──
    multi_op_html = ""
    if len(stats["operators"]) > 1:
        multi_op_html = (
            f'<div class="section">'
            f'<div class="sec-hdr"><div class="sec-ico" style="background:{PRIMARY}1a">'
            f'{ci(ICO_PERSON, PRIMARY)}</div><h2>Показатели операторов</h2></div>'
            f'<table class="dtbl">'
            f'<thead><tr><th>Оператор</th><th class="tc">Звонков</th><th class="tc">Оценка</th>'
            f'<th class="tc">Конверсия</th><th class="tc">Ср. длит.</th><th>Частые ошибки</th></tr></thead>'
            f'<tbody>{op_rows}</tbody></table></div>'
        )

    sc_avg_color = SUCCESS if (stats["avg_score"] or 0) >= 8 else WARNING if (stats["avg_score"] or 0) >= 6 else DANGER

    return f"""<!DOCTYPE html>
<html lang="ru"><head><meta charset="UTF-8">
<style>
@page {{
  size: A4; margin: 14mm 11mm 14mm 11mm;
  @top-center {{ content: "Отчёт колл-центра  {d_from} – {d_to}";
    font-size:8px; color:#94A3B8; font-family:Arial,sans-serif; }}
  @bottom-right {{ content: "стр. " counter(page) " / " counter(pages);
    font-size:8px; color:#94A3B8; font-family:Arial,sans-serif; }}
}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:Arial,"Helvetica Neue",sans-serif;font-size:10px;color:#1E293B;background:#fff}}
/* COVER */
.cover{{background:{PRIMARY};color:#fff;padding:22px 26px 18px;border-radius:8px;margin-bottom:16px}}
.cover-top{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:16px}}
.cover-title{{font-size:18px;font-weight:700;letter-spacing:-.3px;white-space:nowrap}}
.cover-sub{{font-size:11px;opacity:.8;margin-top:3px}}
.cover-meta{{text-align:right;font-size:10px;opacity:.85;line-height:1.9}}
.cover-period{{font-size:13px;font-weight:700;opacity:1}}
.cover-ops{{font-size:9px;opacity:.7;margin-top:2px}}
/* KPI */
.kpi-row{{display:flex;gap:8px;margin-bottom:14px}}
.kpi-card{{flex:1;background:#fff;border:1px solid #E2E8F0;border-radius:7px;padding:10px 8px;text-align:center}}
.kpi-icon{{width:32px;height:32px;border-radius:7px;margin:0 auto 6px;display:flex;align-items:center;justify-content:center}}
.kpi-val{{font-size:18px;font-weight:700;margin-bottom:2px}}
.kpi-lbl{{font-size:8px;text-transform:uppercase;letter-spacing:.4px;color:#64748B}}
.kpi-sub{{font-size:8px;color:#94A3B8;margin-top:1px}}
/* CHARTS */
.charts-row{{display:flex;gap:14px;margin-bottom:14px}}
.chart-box{{flex:1;border:1px solid #E2E8F0;border-radius:7px;padding:12px}}
.chart-ttl{{font-size:10px;font-weight:700;color:#374151;text-transform:uppercase;letter-spacing:.3px;margin-bottom:8px}}
/* SECTIONS */
.section{{margin-bottom:16px}}
.pb{{break-before:page}}
.sec-hdr{{display:flex;align-items:center;gap:8px;margin-bottom:10px;
  border-bottom:2px solid #EEF2FF;padding-bottom:6px}}
.sec-ico{{width:28px;height:28px;border-radius:6px;display:flex;align-items:center;justify-content:center;flex-shrink:0}}
.sec-hdr h2{{font-size:13px;font-weight:700;color:#1E293B}}
.sec-sub{{font-size:9px;color:#94A3B8;margin-left:auto}}
/* TABLE */
.dtbl{{width:100%;border-collapse:collapse;font-size:9px}}
.dtbl thead th{{background:{PRIMARY};color:#fff;padding:7px 6px;text-align:left;
  font-size:9px;font-weight:600;text-transform:uppercase;letter-spacing:.2px}}
.dtbl thead th:first-child{{border-radius:5px 0 0 0}}.dtbl thead th:last-child{{border-radius:0 5px 0 0}}
.dtbl tbody tr:nth-child(even){{background:#F8FAFF}}
.dtbl tbody td{{padding:5px 6px;border-bottom:1px solid #F1F5F9;vertical-align:top}}
.tc{{text-align:center}}.sm{{font-size:8px;color:#64748B}}
.badge{{display:inline-block;padding:2px 7px;border-radius:10px;font-size:8px;font-weight:600;white-space:nowrap}}
/* OPERATOR */
.op-cell{{display:flex;align-items:center;gap:6px}}
.op-av{{width:24px;height:24px;background:{PRIMARY};color:#fff;border-radius:5px;
  display:flex;align-items:center;justify-content:center;font-size:9px;font-weight:700;flex-shrink:0}}
/* DETAIL CARDS */
.det-card{{border:1px solid #E2E8F0;border-radius:6px;padding:10px;margin-bottom:8px;break-inside:avoid}}
.det-hdr{{display:flex;gap:10px;margin-bottom:6px;padding-bottom:5px;border-bottom:1px solid #F1F5F9;font-size:9px}}
.det-date{{color:#64748B}}.det-op{{font-weight:700;color:{PRIMARY}}}.det-svc{{color:#374151}}
.det-sec{{margin-top:5px}}
.det-lbl{{font-size:8px;font-weight:700;text-transform:uppercase;color:#94A3B8;letter-spacing:.3px;margin-bottom:2px}}
.det-txt{{font-size:9px;color:#374151;line-height:1.5}}
.err-ul{{list-style:none}}.err-ul li{{font-size:9px;color:#991B1B;padding:1px 0}}
.err-ul li::before{{content:"• ";color:#DC2626}}
/* RECOMMENDATIONS */
.recs-col{{display:flex;flex-direction:column;gap:8px}}
.rec-card{{border-left:3px solid {PRIMARY};padding:9px 12px;background:#EEF2FF;border-radius:0 6px 6px 0;break-inside:avoid}}
.rec-title{{font-size:11px;font-weight:700;color:{PRIMARY};margin-bottom:3px}}
.rec-txt{{font-size:9px;color:#374151;line-height:1.5}}
</style></head><body>

<div class="cover">
  <div class="cover-top">
    <div>
      <div class="cover-title">Аналитический отчёт</div>
    </div>
    <div class="cover-meta">
      <div class="cover-period">{d_from} &mdash; {d_to}</div>
      <div>Сформирован: {generated}</div>
      <div class="cover-ops">{ops_label}</div>
    </div>
  </div>
</div>

<div class="kpi-row">
  {kpi(ICO_PHONE, "Всего звонков",  stats["total"], None, PRIMARY)}
  {kpi(ICO_CHECK, "Конверсия",      f'{stats["conversion"]}%', f'{stats["recorded"]} записей',
       SUCCESS if stats["conversion"] >= 50 else WARNING)}
  {kpi(ICO_STAR,  "Средняя оценка", stats["avg_score"] if stats["avg_score"] is not None else "—", "из 10", sc_avg_color)}
  {kpi(ICO_CLOCK, "Ср. длительность", _fmt_dur(stats["avg_dur"]), None, PRIMARY)}
  {kpi(ICO_PERSON,"Операторов",     len(stats["operators"]), None, "#7C3AED")}
</div>

<div class="charts-row">
  <div class="chart-box">
    <div class="chart-ttl">Итоги звонков</div>{outcome_svg}
  </div>
  <div class="chart-box">
    <div class="chart-ttl">Распределение оценок</div>{score_svg}
  </div>
</div>

{multi_op_html}

<div class="section">
  <div class="sec-hdr">
    <div class="sec-ico" style="background:{PRIMARY}1a">{ci(ICO_LIST, PRIMARY)}</div>
    <h2>Детализация звонков</h2>
    <span class="sec-sub">Всего: {stats["total"]}</span>
  </div>
  <table class="dtbl">
  <thead><tr>
    <th>Дата / Время</th><th>Оператор</th><th class="tc">Длит.</th>
    <th>Итог</th><th>Тема / Услуга</th><th class="tc">Оценка</th><th>Тон</th><th>Качество</th>
  </tr></thead>
  <tbody>{call_rows}</tbody>
  </table>
</div>

{summaries_html}
{ai_html}
</body></html>"""


@app.route("/api/report/operators")
@login_required
def api_report_operators():
    uid = session["user_id"]
    ops = db.get_operators_list(uid)
    with db.get_conn() as conn:
        row = conn.execute(
            "SELECT COUNT(*) FROM calls WHERE user_id=? AND (operator IS NULL OR operator='')",
            (uid,),
        ).fetchone()
        has_unknown = (row[0] > 0)
    return jsonify({"operators": ops, "has_unknown": has_unknown})


@app.route("/api/report/pdf", methods=["POST"])
@login_required
def api_report_pdf():
    data = request.get_json() or {}
    date_from         = data.get("date_from")
    date_to           = data.get("date_to")
    operators         = data.get("operators") or []
    ai_recommendations = data.get("ai_recommendations", False)
    include_summaries  = data.get("include_summaries", True)

    user_id = session["user_id"]

    if operators:
        calls = []
        for op in operators:
            calls += db.get_all_calls(user_id, date_from=date_from, date_to=date_to, operator=op)
        seen = set()
        calls = [c for c in calls if not (c["id"] in seen or seen.add(c["id"]))]
    else:
        calls = db.get_all_calls(user_id, date_from=date_from, date_to=date_to)

    if not calls:
        return jsonify({"error": "Нет данных за выбранный период"}), 404

    try:
        stats   = _calc_report_stats(calls)
        gc_token = db.get_user_gigachat_token(user_id) or None
        ai_recs = _generate_ai_recommendations(calls, gc_token) if ai_recommendations else []
        html    = _build_report_html(calls, stats, ai_recs, date_from, date_to, operators, include_summaries)
        from weasyprint import HTML as WP
        pdf_bytes = WP(string=html).write_pdf()
    except Exception as e:
        import traceback
        print(f"[report] error: {traceback.format_exc()}", flush=True)
        return jsonify({"error": f"Ошибка генерации: {e}"}), 500

    fname = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── Audio / Transcript ───────────────────────────────────────────────────────

@app.route("/api/calls/<int:call_id>/audio")
@login_required
def api_audio(call_id):
    row = db.get_call(call_id, session["user_id"])
    if not row:
        return jsonify({"error": "not found"}), 404
    path = os.path.join(AUDIO_DIR, row["filename"])
    if not os.path.exists(path):
        return jsonify({"error": "file not found"}), 404
    return send_file(path, mimetype="audio/mpeg", conditional=True)


@app.route("/api/calls/<int:call_id>/transcript")
@login_required
def api_transcript(call_id):
    row = db.get_call(call_id, session["user_id"])
    if not row or not row.get("transcript"):
        return jsonify({"error": "not found"}), 404

    errors = row.get("operator_errors") or []
    if isinstance(errors, str):
        try:
            errors = json.loads(errors)
        except Exception:
            errors = [errors]

    lines = [
        f"Файл: {row['filename']}",
        f"Дата звонка: {row.get('call_date', '—')} {row.get('call_time', '')}",
        f"Оператор: {row.get('operator', '—')}",
        f"Итог: {row.get('result', '—')}",
        f"Тема: {row.get('service', '—')}",
        f"Оценка оператора: {row.get('score', '—')}/10",
        f"Тон: {row.get('tone', '—')}",
        f"Качество: {row.get('call_quality', '—')}",
        "",
        "─" * 60,
        "ТРАНСКРИПТ РАЗГОВОРА",
        "─" * 60,
        "",
        row["transcript"],
        "",
    ]
    if row.get("summary"):
        lines += ["─" * 60, "РЕЗЮМЕ", "─" * 60, "", row["summary"], ""]
    if errors:
        lines += ["─" * 60, "ОШИБКИ ОПЕРАТОРА", "─" * 60]
        for e in errors:
            lines.append(f"  * {e}")
        lines.append("")
    if row.get("incorrect_info"):
        lines += ["─" * 60, "НЕКОРРЕКТНОЕ ИНФОРМИРОВАНИЕ", "─" * 60,
                  "", row["incorrect_info"], ""]
    if row.get("recommendations"):
        lines += ["─" * 60, "РЕКОМЕНДАЦИИ", "─" * 60, "", row["recommendations"], ""]

    text = "\n".join(lines)
    safe_name = os.path.splitext(row["filename"])[0]
    return Response(
        text,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="transcript_{safe_name}.txt"'}
    )


# ─── App settings ─────────────────────────────────────────────────────────────

@app.route("/api/settings/config")
@login_required
def api_settings_get():
    return jsonify(db.get_all_settings())


@app.route("/api/settings/config", methods=["POST"])
@login_required
@admin_required
def api_settings_set():
    data = request.json or {}
    allowed_keys = {"conversion_mode", "conversion_enabled", "quality_enabled"}
    for key, val in data.items():
        if key in allowed_keys:
            db.set_setting(key, str(val))
    return jsonify({"ok": True})


# ─── Tags catalog ─────────────────────────────────────────────────────────────

@app.route("/api/tags")
@login_required
def api_tags_list():
    return jsonify(db.get_tags_catalog())


@app.route("/api/tags", methods=["POST"])
@login_required
def api_tags_create():
    data = request.json or {}
    name = data.get("name", "").strip()
    color = data.get("color", "#6366F1")
    if not name:
        return jsonify({"error": "name required"}), 400
    tag_id = db.create_tag(name, color, session["user_id"])
    if not tag_id:
        return jsonify({"error": "Тег уже существует"}), 409
    return jsonify({"ok": True, "id": tag_id})


@app.route("/api/tags/<int:tag_id>", methods=["DELETE"])
@login_required
def api_tags_delete(tag_id):
    db.delete_tag(tag_id)
    return jsonify({"ok": True})


# ─── Settings: Price files (per-user) ─────────────────────────────────────────

@app.route("/api/settings/prices")
@login_required
def api_price_list():
    return jsonify(db.get_price_files(session["user_id"]))


@app.route("/api/settings/prices", methods=["POST"])
@login_required
def api_price_upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400
    if not f.filename.lower().endswith(".txt"):
        return jsonify({"error": "только .txt файлы"}), 400
    display_name = secure_filename(f.filename)
    import uuid as _uuid
    stored_name = f"{_uuid.uuid4().hex}_{display_name}"
    os.makedirs(PRICE_DIR, exist_ok=True)
    f.save(os.path.join(PRICE_DIR, stored_name))
    price_id = db.add_price_file(stored_name, display_name, session["user_id"])
    return jsonify({"ok": True, "id": price_id})


@app.route("/api/settings/prices/<int:price_id>/activate", methods=["POST"])
@login_required
def api_price_activate(price_id):
    db.activate_price(price_id, session["user_id"])
    import prices as _prices
    _prices.invalidate_cache(session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/settings/prices/<int:price_id>/deactivate", methods=["POST"])
@login_required
def api_price_deactivate(price_id):
    db.deactivate_price(session["user_id"])
    import prices as _prices
    _prices.invalidate_cache(session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/settings/prices/<int:price_id>", methods=["DELETE"])
@login_required
def api_price_delete(price_id):
    filename = db.delete_price_file(price_id, session["user_id"])
    if not filename:
        return jsonify({"error": "not found"}), 404
    path = os.path.join(PRICE_DIR, filename)
    if os.path.exists(path):
        os.remove(path)
    import prices as _prices
    _prices.invalidate_cache(session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/settings/prices/<int:price_id>/content")
@login_required
def api_price_content_get(price_id):
    content = db.get_price_content(price_id)
    if content is None:
        return jsonify({"error": "not found"}), 404
    return jsonify({"content": content})


@app.route("/api/settings/prices/<int:price_id>/content", methods=["PUT"])
@login_required
def api_price_content_update(price_id):
    data = request.json or {}
    content = data.get("content", "")
    if not content.strip():
        return jsonify({"error": "Содержимое не может быть пустым"}), 400
    db.update_price_content(price_id, content)
    import prices as _prices
    _prices.invalidate_cache(session["user_id"])
    return jsonify({"ok": True})


# ─── Settings: Prompt files (per-user + универсальные) ────────────────────────

@app.route("/api/settings/prompts")
@login_required
def api_prompt_list():
    return jsonify(db.get_prompt_files(session["user_id"]))


@app.route("/api/settings/prompts", methods=["POST"])
@login_required
def api_prompt_upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400
    if not f.filename.lower().endswith(".txt"):
        return jsonify({"error": "только .txt файлы"}), 400
    display_name = secure_filename(f.filename)
    content = f.read().decode("utf-8", errors="replace").strip()
    if not content:
        return jsonify({"error": "файл пустой"}), 400
    prompt_id = db.add_prompt(display_name, content, session["user_id"])
    return jsonify({"ok": True, "id": prompt_id})


@app.route("/api/settings/prompts/<int:prompt_id>/activate", methods=["POST"])
@login_required
def api_prompt_activate(prompt_id):
    db.activate_prompt(prompt_id, session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/settings/prompts/<int:prompt_id>/deactivate", methods=["POST"])
@login_required
def api_prompt_deactivate(prompt_id):
    db.deactivate_prompt(session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/settings/prompts/<int:prompt_id>", methods=["DELETE"])
@login_required
def api_prompt_delete(prompt_id):
    db.delete_prompt(prompt_id, session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/settings/prompts/<int:prompt_id>/content")
@login_required
def api_prompt_content_get(prompt_id):
    content = db.get_prompt_content(prompt_id)
    if content is None:
        return jsonify({"error": "not found"}), 404
    return jsonify({"content": content})


@app.route("/api/settings/prompts/<int:prompt_id>/content", methods=["PUT"])
@login_required
def api_prompt_content_update(prompt_id):
    data = request.json or {}
    content = data.get("content", "").strip()
    if not content:
        return jsonify({"error": "Содержимое не может быть пустым"}), 400
    db.update_prompt_content(prompt_id, content)
    return jsonify({"ok": True})


@app.route("/api/settings/default-prompt")
@login_required
def api_default_prompt():
    from analyzer import DEFAULT_SYSTEM_PROMPT
    return jsonify({"content": DEFAULT_SYSTEM_PROMPT})


if __name__ == "__main__":
    db.init_db()
    os.makedirs(AUDIO_DIR, exist_ok=True)
    scan_and_process()
    app.run(host="0.0.0.0", port=PORT)
